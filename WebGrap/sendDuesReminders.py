import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duessdf.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
lastesMonth = sheet.cell(row = 1, column = lastCol).value
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row = r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column = 1).value
        email = sheet.cell(row=r,column=2).value
        unpaidMembers[name] = email
smtpobj = smtplib.SMTP('smtp.gmail.com',587)   
smtpobj.ehlo()
smtpobj.starttls()
smtpobj.login('my_email_address@gmail.com', sys.argv[1])
for name, email in unpaidMembers.items():
     body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not
paid dues for %s. Please make this payment as soon as possible. Thank you!'" %
(latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)
    if sendmailStatus != {}:
          print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()